from __future__ import annotations

import io
import csv
import re
import unicodedata
from dataclasses import dataclass
from functools import lru_cache
from datetime import datetime
from typing import BinaryIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ACCOUNT_RE = re.compile(r"^\s*(\d+)\s*-\s*(.+?)\s*$")
DATE_RE = re.compile(r"^\s*(\d{2}/\d{2}/\d{4})\s*$")

REQUIRED_LEDGER_COLUMNS = [
    "Hist\u00f3rico",
    "Chave",
    "Contra",
    "D\u00e9bito",
    "Cr\u00e9dito",
    "Saldo",
]

VALUE_LEDGER_COLUMNS = [
    "Hist\u00f3rico",
    "Chave",
    "Contra",
    "Valor",
    "Saldo",
]

PLAN_COLUMNS = [
    "C\u00f3digo",
    "Classifica\u00e7\u00e3o",
    "Tipo",
    "Nome",
    "Grupo",
    "Relat\u00f3rio",
    "Saldo",
]

REDUCER_TERMS = [
    "(-)",
    "( - )",
    "redutora",
    "deprecia\u00e7\u00e3o acumulada",
    "amortiza\u00e7\u00e3o acumulada",
    "exaust\u00e3o acumulada",
    "perdas estimadas",
    "provis\u00e3o para perdas",
    "ajuste a valor presente",
    "juros a apropriar",
    "encargos a transcorrer",
    "duplicatas descontadas",
    "capital a integralizar",
]


@dataclass
class LedgerEntry:
    codigo: str
    nome_razao: str
    data: pd.Timestamp
    debito: float
    credito: float
    saldo_final_dia: float
    lado_saldo: str


@lru_cache(maxsize=20000)
def normalize_text_cached(text: str) -> str:
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", text).strip().lower()


def normalize_text(value: object) -> str:
    text = "" if pd.isna(value) else str(value)
    return normalize_text_cached(text)


def normalize_code(value: object) -> str:
    if pd.isna(value):
        return ""

    text = str(value).strip()
    if not text:
        return ""

    if re.fullmatch(r"\d+([,.]0+)?", text):
        text = re.split(r"[,.]", text)[0]

    return text.strip()


def is_continuation_header(value: object) -> bool:
    return "continuacao" in normalize_text(value)


@lru_cache(maxsize=50000)
def parse_brazilian_number_cached(text: str) -> float:
    text = text.strip()
    if not text:
        return 0.0

    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()")
    text = re.sub(r"[^\d,.\-]", "", text)

    if not text:
        return 0.0

    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    else:
        pieces = text.split(".")
        if len(pieces) > 2:
            text = "".join(pieces)

    try:
        number = float(text)
    except ValueError:
        return 0.0

    return -abs(number) if negative else number


def parse_brazilian_number(value: object) -> float:
    if pd.isna(value):
        return 0.0
    return parse_brazilian_number_cached(str(value))


def parse_balance_value(value: object) -> tuple[float, str]:
    if pd.isna(value):
        return 0.0, ""

    text = str(value).strip()
    side = ""
    normalized = normalize_text(text)
    if re.search(r"\d\s*d$", normalized) or re.search(r"(^|[\s\-/])d($|[\s\-/])", normalized):
        side = "D"
    elif re.search(r"\d\s*c$", normalized) or re.search(r"(^|[\s\-/])c($|[\s\-/])", normalized):
        side = "C"

    return parse_brazilian_number(text), side


@lru_cache(maxsize=5000)
def parse_date_cached(text: str) -> pd.Timestamp:
    try:
        return pd.Timestamp(datetime.strptime(text, "%d/%m/%Y"))
    except ValueError:
        return pd.NaT


def read_csv_semicolon(uploaded_file: BinaryIO) -> pd.DataFrame:
    raw = uploaded_file.read()
    try:
        uploaded_file.seek(0)
    except (AttributeError, OSError):
        pass

    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "latin1", "cp1252"):
        try:
            return pd.read_csv(
                io.BytesIO(raw),
                sep=";",
                dtype=str,
                encoding=encoding,
                keep_default_na=False,
            )
        except UnicodeDecodeError as exc:
            last_error = exc
            continue
        except pd.errors.ParserError:
            try:
                return read_csv_semicolon_relaxed(raw, encoding)
            except Exception as exc:
                last_error = exc
                continue

    if last_error:
        raise last_error
    return read_csv_semicolon_relaxed(raw, "latin1")


def read_csv_semicolon_relaxed(raw: bytes, encoding: str) -> pd.DataFrame:
    text = raw.decode(encoding)
    rows = list(csv.reader(io.StringIO(text), delimiter=";"))
    if not rows:
        return pd.DataFrame()

    header = rows[0]
    width = len(header)
    is_ledger = [normalize_text(column) for column in header] == [
        normalize_text(column) for column in REQUIRED_LEDGER_COLUMNS
    ]

    fixed_rows = []
    for row in rows[1:]:
        if len(row) < width:
            row = row + [""] * (width - len(row))
        elif len(row) > width:
            extra = len(row) - width
            if is_ledger:
                row = [";".join(row[: extra + 1])] + row[extra + 1 :]
            else:
                row = row[: width - 1] + [";".join(row[width - 1 :])]
        fixed_rows.append(row)

    return pd.DataFrame(fixed_rows, columns=header).fillna("")


def rename_columns(df: pd.DataFrame, expected: list[str]) -> pd.DataFrame:
    lookup = {normalize_text(column): column for column in df.columns}
    rename_map = {}

    for column in expected:
        actual = lookup.get(normalize_text(column))
        if actual:
            rename_map[actual] = column

    return df.rename(columns=rename_map)


def validate_columns(df: pd.DataFrame, required: list[str], file_label: str) -> list[str]:
    missing = [column for column in required if column not in df.columns]
    if missing:
        return [f"{file_label}: coluna ausente: {', '.join(missing)}"]
    return []


def normalize_ledger_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = rename_columns(df, REQUIRED_LEDGER_COLUMNS)
    df = rename_columns(df, VALUE_LEDGER_COLUMNS)

    if all(column in df.columns for column in REQUIRED_LEDGER_COLUMNS):
        return df

    if all(column in df.columns for column in VALUE_LEDGER_COLUMNS):
        df = df.copy()
        debits = [""] * len(df)
        credits = [""] * len(df)
        current_code = ""
        current_name = ""
        current_date: pd.Timestamp | None = None
        movements: list[dict[str, object]] = []

        def flush_day() -> None:
            if not movements:
                return

            total_debit = None
            total_credit = None
            for movement in movements:
                hist_norm = normalize_text(movement["historico"])
                if "total dia" in hist_norm:
                    total_debit = parse_brazilian_number(movement["chave"])
                    total_credit = parse_brazilian_number(str(movement["historico"]).split("credito:")[-1] if "credito:" in hist_norm else movement["valor"])
                    break

            regular = [movement for movement in movements if "total dia" not in normalize_text(movement["historico"])]
            if not regular:
                return

            values = [parse_brazilian_number(movement["valor"]) for movement in regular]
            sum_values = round(sum(values), 2)

            if total_debit is not None and abs(sum_values - total_debit) <= 0.02 and abs((total_credit or 0.0)) <= 0.02:
                for movement, value in zip(regular, values):
                    debits[int(movement["index"])] = format_brazilian_number(value)
                return

            if total_credit is not None and abs(sum_values - total_credit) <= 0.02 and abs((total_debit or 0.0)) <= 0.02:
                for movement, value in zip(regular, values):
                    credits[int(movement["index"])] = format_brazilian_number(value)
                return

            previous = None
            for movement, value in zip(regular, values):
                saldo = parse_brazilian_number(movement["saldo"])
                if previous is None:
                    delta = saldo - parse_brazilian_number(movement["saldo_anterior"])
                else:
                    delta = saldo - previous
                previous = saldo
                if delta >= 0:
                    credits[int(movement["index"])] = format_brazilian_number(value)
                else:
                    debits[int(movement["index"])] = format_brazilian_number(value)

        historicos = df["Hist\u00f3rico"].astype(str).tolist()
        chaves = df["Chave"].astype(str).tolist()
        contras = df["Contra"].astype(str).tolist()
        valores = df["Valor"].astype(str).tolist()
        saldos = df["Saldo"].astype(str).tolist()

        for index, (historico_raw, chave, contra, valor_raw, saldo) in enumerate(
            zip(historicos, chaves, contras, valores, saldos)
        ):
            historico = historico_raw.strip()
            valor_text = valor_raw.strip()
            hist_norm = normalize_text(historico)
            account_match = ACCOUNT_RE.match(historico)
            is_header = (
                account_match
                and not chave.strip()
                and not contra.strip()
                and (is_blank(valor_text) or normalize_text(valor_text).startswith("saldo anterior") or normalize_text(valor_text).startswith("saldo da pagina anterior"))
            )
            if is_header:
                next_code = normalize_code(account_match.group(1))
                next_name = clean_account_name(account_match.group(2))
                if (
                    is_continuation_header(historico)
                    and current_code == next_code
                    and current_name == next_name
                    and current_date is not None
                    and not pd.isna(current_date)
                ):
                    continue

                flush_day()
                movements = []
                current_code = next_code
                current_name = next_name
                current_date = None
                continue

            date_match = DATE_RE.match(historico)
            if date_match:
                flush_day()
                movements = []
                current_date = parse_date_cached(date_match.group(1))
                continue

            if not current_code or current_date is None or pd.isna(current_date):
                continue

            if not valor_text:
                continue

            movements.append(
                {
                    "index": index,
                    "historico": historico,
                    "chave": chave,
                    "valor": valor_text,
                    "saldo": saldo,
                    "saldo_anterior": saldos[index - 1] if index > 0 else "0",
                }
            )

        flush_day()

        df["D\u00e9bito"] = debits
        df["Cr\u00e9dito"] = credits

        return df[REQUIRED_LEDGER_COLUMNS]

    return df


def format_brazilian_number(value: float) -> str:
    if not value:
        return ""
    return f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def is_blank(value: object) -> bool:
    return pd.isna(value) or str(value).strip() == ""


def code_sort_key(value: str) -> tuple[int, object]:
    return (0, int(value)) if value.isdigit() else (1, value)


def clean_account_name(name: str) -> str:
    text = name.strip()
    if normalize_text(text).endswith("(continuacao)"):
        return re.sub(r"\s*\([^)]*\)\s*$", "", text).strip()
    return text


def parse_ledger(df: pd.DataFrame) -> pd.DataFrame:
    df = normalize_ledger_columns(df)
    errors = validate_columns(df, REQUIRED_LEDGER_COLUMNS, "Razao")
    if errors:
        raise ValueError("\n".join(errors))

    current_code = ""
    current_name = ""
    current_date: pd.Timestamp | None = None
    daily: dict[tuple[str, str, pd.Timestamp], LedgerEntry] = {}

    historicos = df["Hist\u00f3rico"].astype(str).tolist()
    chaves = df["Chave"].astype(str).tolist()
    contras = df["Contra"].astype(str).tolist()
    debitos = df["D\u00e9bito"].astype(str).tolist()
    creditos = df["Cr\u00e9dito"].astype(str).tolist()
    saldos = df["Saldo"].astype(str).tolist()

    for historico_raw, chave, contra, debito_raw, credito_raw, saldo_raw in zip(
        historicos, chaves, contras, debitos, creditos, saldos
    ):
        historico = historico_raw.strip()
        debito_text = debito_raw.strip()

        account_match = ACCOUNT_RE.match(historico)
        debito_norm = normalize_text(debito_text)
        is_account_header = (
            account_match
            and not chave.strip()
            and not contra.strip()
            and not credito_raw.strip()
            and (
                is_blank(debito_text)
                or debito_norm.startswith("saldo anterior")
                or debito_norm.startswith("saldo da pagina anterior")
            )
        )
        if is_account_header:
            next_code = normalize_code(account_match.group(1))
            next_name = clean_account_name(account_match.group(2))
            is_same_continuation = (
                is_continuation_header(historico)
                and current_code == next_code
                and current_name == next_name
                and current_date is not None
                and not pd.isna(current_date)
            )
            current_code = next_code
            current_name = next_name
            if not is_same_continuation and not debito_norm.startswith("saldo da pagina anterior"):
                current_date = None
            continue

        date_match = DATE_RE.match(historico)
        if date_match:
            current_date = parse_date_cached(date_match.group(1))
            continue

        if not current_code or current_date is None or pd.isna(current_date):
            continue

        saldo_text = saldo_raw.strip()
        if not saldo_text:
            continue

        saldo_value, saldo_side = parse_balance_value(saldo_text)
        key = (current_code, current_name, current_date)
        debito = parse_brazilian_number(debito_raw)
        credito = parse_brazilian_number(credito_raw)
        if key in daily:
            entry = daily[key]
            daily[key] = LedgerEntry(
                codigo=entry.codigo,
                nome_razao=entry.nome_razao,
                data=entry.data,
                debito=entry.debito + debito,
                credito=entry.credito + credito,
                saldo_final_dia=saldo_value,
                lado_saldo=saldo_side,
            )
        else:
            daily[key] = LedgerEntry(
                codigo=current_code,
                nome_razao=current_name,
                data=current_date,
                debito=debito,
                credito=credito,
                saldo_final_dia=saldo_value,
                lado_saldo=saldo_side,
            )

    return pd.DataFrame([entry.__dict__ for entry in daily.values()])


def ledger_file_diagnostics(df: pd.DataFrame) -> dict[str, object]:
    df = rename_columns(df, REQUIRED_LEDGER_COLUMNS)
    df = rename_columns(df, VALUE_LEDGER_COLUMNS)

    has_standard = all(column in df.columns for column in REQUIRED_LEDGER_COLUMNS)
    has_value = all(column in df.columns for column in VALUE_LEDGER_COLUMNS)
    if not has_standard and not has_value:
        return {"account_codes": []}

    account_codes: set[str] = set()
    historicos = df["Hist\u00f3rico"].astype(str).tolist()
    chaves = df["Chave"].astype(str).tolist()
    contras = df["Contra"].astype(str).tolist()
    debit_or_value_column = "D\u00e9bito" if has_standard else "Valor"
    debit_or_values = df[debit_or_value_column].astype(str).tolist()
    credits = df["Cr\u00e9dito"].astype(str).tolist() if has_standard else [""] * len(df)

    for historico_raw, chave, contra, debit_or_value, credit in zip(
        historicos, chaves, contras, debit_or_values, credits
    ):
        historico = historico_raw.strip()
        account_match = ACCOUNT_RE.match(historico)
        if account_match:
            debit_text = debit_or_value.strip()
            debit_norm = normalize_text(debit_text)
            is_header = (
                not chave.strip()
                and not contra.strip()
                and not credit.strip()
                and (
                    is_blank(debit_text)
                    or debit_norm.startswith("saldo anterior")
                    or debit_norm.startswith("saldo da pagina anterior")
                )
            )
            if is_header:
                account_codes.add(normalize_code(account_match.group(1)))

    return {
        "account_codes": sorted(account_codes, key=code_sort_key),
    }


def prepare_plan(df: pd.DataFrame) -> pd.DataFrame:
    df = rename_columns(df, PLAN_COLUMNS)
    errors = validate_columns(df, ["C\u00f3digo"], "Plano de contas")
    if errors:
        raise ValueError("\n".join(errors))

    for column in PLAN_COLUMNS:
        if column not in df.columns:
            df[column] = ""

    df = df[PLAN_COLUMNS].copy()
    df["codigo_normalizado"] = df["C\u00f3digo"].map(normalize_code)
    return df.drop_duplicates(subset=["codigo_normalizado"], keep="first")


def infer_base_nature(row: pd.Series) -> tuple[str, str]:
    combined = normalize_text(
        " ".join(
            str(row.get(column, ""))
            for column in ["Classifica\u00e7\u00e3o", "Tipo", "Nome", "Grupo", "Relat\u00f3rio", "Saldo"]
        )
    )

    if any(term in combined for term in ["compensacao", "compensado", "compensadas"]):
        return "revisao", "Conta de compensacao: revisar manualmente."

    debit_terms = ["devedor", "devedora", "ativo", "despesa", "despesas", "custo", "custos"]
    credit_terms = ["credor", "credora", "passivo", "patrimonio liquido", "receita", "receitas"]

    if any(term in combined for term in debit_terms):
        return "devedora", ""
    if any(term in combined for term in credit_terms):
        return "credora", ""

    first_classification = str(row.get("Classifica\u00e7\u00e3o", "")).strip()[:1]
    if first_classification in {"1", "4", "5"}:
        return "devedora", ""
    if first_classification in {"2", "3"}:
        return "credora", ""

    return "indefinida", "Natureza nao identificada pelo plano de contas."


def is_reducer(row: pd.Series) -> bool:
    combined = normalize_text(
        " ".join(str(row.get(column, "")) for column in ["nome_razao", "Nome", "Grupo", "Classifica\u00e7\u00e3o"])
    )
    normalized_terms = [normalize_text(term) for term in REDUCER_TERMS]
    return any(term in combined for term in normalized_terms)


def invert_nature(nature: str) -> str:
    if nature == "devedora":
        return "credora"
    if nature == "credora":
        return "devedora"
    return nature


def is_participant_account_name(value: object) -> bool:
    text = "" if pd.isna(value) else str(value).strip()
    return bool(re.match(r"^\d+\s+-\s+", text))


def movement_impact(row: pd.Series) -> float:
    if row.get("Natureza esperada") == "credora":
        return float(row.get("credito", 0.0)) - float(row.get("debito", 0.0))
    if row.get("Natureza esperada") == "devedora":
        return float(row.get("debito", 0.0)) - float(row.get("credito", 0.0))
    return 0.0


def recalculate_running_balances(result: pd.DataFrame) -> pd.DataFrame:
    result = result.copy()
    result["_ordem_original"] = range(len(result))
    result["_impacto"] = result.apply(movement_impact, axis=1)

    recalculated: list[pd.Series] = []
    for _, group in result.groupby(["codigo", "nome_razao"], sort=False):
        group = group.sort_values(["data", "_ordem_original"]).copy()
        if group.empty:
            continue

        first = group.iloc[0]
        initial_balance = float(first["saldo_final_dia"]) - float(first["_impacto"])
        running_balance = initial_balance

        for index, row in group.iterrows():
            running_balance += float(row["_impacto"])
            result.at[index, "saldo_final_dia"] = round(running_balance, 2)
            result.at[index, "Saldo recalculado por movimentos"] = "sim"

    return result.drop(columns=["_ordem_original", "_impacto"])


def collapse_issue_sequences(output: pd.DataFrame) -> pd.DataFrame:
    if output[output["Tipo de inconsistencia"].ne("")].empty:
        return output[output["Tipo de inconsistencia"].ne("")].copy()

    work = output.copy()
    work["_data_dt"] = pd.to_datetime(work["Data"], format="%d/%m/%Y", errors="coerce")
    kept_rows = []

    for _, group in work.sort_values(["Conta analisada", "_data_dt"]).groupby(
        ["Codigo da conta", "Conta analisada"],
        sort=False,
    ):
        sequence_rows = []
        sequence_type = ""

        def flush_sequence(rows: list[pd.Series]) -> None:
            if not rows:
                return

            first = rows[0].copy()
            first["Dias impactados"] = len(rows)
            first["Data final da sequencia"] = rows[-1]["Data"]
            if len(rows) > 1:
                extra = (
                    f" Sequencia negativa resumida: {len(rows)} dias impactados, "
                    f"de {first['Data']} ate {rows[-1]['Data']}."
                )
                first["Observacao"] = (str(first.get("Observacao", "")).strip() + extra).strip()
            kept_rows.append(first)

        for _, row in group.iterrows():
            issue_type = str(row.get("Tipo de inconsistencia", "")).strip()
            if not issue_type:
                flush_sequence(sequence_rows)
                sequence_rows = []
                sequence_type = ""
                continue

            if sequence_rows and issue_type != sequence_type:
                flush_sequence(sequence_rows)
                sequence_rows = []

            sequence_type = issue_type
            sequence_rows.append(row)

        flush_sequence(sequence_rows)

    collapsed = pd.DataFrame(kept_rows).drop(columns=["_data_dt"], errors="ignore")
    return collapsed.reset_index(drop=True)


def analyze_balances(ledger_df: pd.DataFrame, plan_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    plan = prepare_plan(plan_df)
    ledger = parse_ledger(ledger_df)

    if ledger.empty:
        raise ValueError("Nenhum saldo diario foi encontrado no razao informado.")

    result = ledger.merge(
        plan,
        left_on="codigo",
        right_on="codigo_normalizado",
        how="left",
        suffixes=("", "_plano"),
    )

    result["Nome no plano de contas"] = result["Nome"].fillna("")
    result["Conta encontrada no plano"] = result["codigo_normalizado"].fillna("").ne("")
    result["Base da natureza"] = result.apply(lambda row: infer_base_nature(row)[0], axis=1)
    result["Observacao"] = result.apply(lambda row: infer_base_nature(row)[1], axis=1)
    result["Eh redutora"] = result.apply(is_reducer, axis=1)
    result["Natureza esperada"] = result.apply(
        lambda row: invert_nature(row["Base da natureza"]) if row["Eh redutora"] else row["Base da natureza"],
        axis=1,
    )

    missing_plan = ~result["Conta encontrada no plano"]
    result.loc[missing_plan, "Observacao"] = "Conta nao encontrada no plano de contas."
    result.loc[missing_plan, "Natureza esperada"] = "indefinida"

    result["Saldo recalculado por movimentos"] = "nao"
    participant_account = result["nome_razao"].map(is_participant_account_name)
    recalculable = result["Natureza esperada"].isin(["devedora", "credora"]) & participant_account
    if recalculable.any():
        recalculated = recalculate_running_balances(result.loc[recalculable].copy())
        result.loc[recalculated.index, ["saldo_final_dia", "Saldo recalculado por movimentos"]] = recalculated[
            ["saldo_final_dia", "Saldo recalculado por movimentos"]
        ]

    result["Tipo de inconsistencia"] = ""
    normal_negative_balance = (~result["Eh redutora"]) & (result["saldo_final_dia"] < 0)
    reducer_positive_balance = result["Eh redutora"] & (result["saldo_final_dia"] > 0)
    compensation = result["Natureza esperada"].eq("revisao")
    undefined = result["Natureza esperada"].eq("indefinida")

    result.loc[normal_negative_balance, "Tipo de inconsistencia"] = "Saldo negativo no razao SCI"
    result.loc[
        normal_negative_balance & result["Natureza esperada"].eq("credora"),
        "Tipo de inconsistencia",
    ] = "Saldo devedor em conta de natureza credora"
    result.loc[
        normal_negative_balance & result["Natureza esperada"].eq("devedora"),
        "Tipo de inconsistencia",
    ] = "Saldo credor em conta de natureza devedora"
    result.loc[reducer_positive_balance, "Tipo de inconsistencia"] = "Conta redutora com saldo positivo no razao SCI"
    result.loc[compensation, "Tipo de inconsistencia"] = "Conta de compensacao para revisao"
    result.loc[undefined, "Tipo de inconsistencia"] = "Natureza nao identificada"

    needs_negative_review = normal_negative_balance & result["Observacao"].fillna("").eq("")
    result.loc[needs_negative_review, "Observacao"] = (
        "O SCI exibiu saldo negativo para esta conta/data. Conferir se o saldo esta invertido."
    )
    needs_reducer_review = reducer_positive_balance & result["Observacao"].fillna("").eq("")
    result.loc[needs_reducer_review, "Observacao"] = (
        "Conta redutora costuma aparecer negativa no SCI. Conferir saldo positivo nesta data."
    )
    output = pd.DataFrame(
        {
            "Codigo da conta": result["codigo"],
            "Conta analisada": result["codigo"] + " - " + result["nome_razao"],
            "Nome da conta no razao": result["nome_razao"],
            "Nome no plano de contas": result["Nome no plano de contas"],
            "Classificacao": result["Classifica\u00e7\u00e3o"].fillna(""),
            "Grupo": result["Grupo"].fillna(""),
            "Natureza esperada": result["Natureza esperada"],
            "Se e redutora": result["Eh redutora"].map({True: "sim", False: "nao"}),
            "Data": result["data"].dt.strftime("%d/%m/%Y"),
            "Saldo final do dia": result["saldo_final_dia"],
            "Lado do saldo": result["lado_saldo"],
            "Tipo de inconsistencia": result["Tipo de inconsistencia"],
            "Observacao": result["Observacao"].fillna(""),
            "Dias impactados": "",
            "Data final da sequencia": "",
        }
    )

    inconsistencies = collapse_issue_sequences(output)
    return output, inconsistencies


AZUL_ESCURO = "1F3864"
AZUL_MED = "2E5FA3"
AZUL_CLARO = "D6E4F0"
VERMELHO_BG = "FCE4D6"
AMARELO_BG = "FFF2CC"
VERDE_BG = "E2EFDA"
CINZA_LINHA = "F2F2F2"
BRANCO = "FFFFFF"
VERMELHO_FG = "C00000"
AMARELO_FG = "7F6000"


def borda_fina():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def borda_media():
    s = Side(style="medium", color=AZUL_ESCURO)
    return Border(left=s, right=s, top=s, bottom=s)


def estilo_cabecalho(cell, bg=AZUL_ESCURO, fg=BRANCO):
    cell.font = Font(name="Arial", bold=True, color=fg, size=9)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = borda_fina()


def issue_fill(value: object) -> str:
    text = normalize_text(value)
    if "credor em conta de natureza devedora" in text:
        return VERMELHO_BG
    if "devedor em conta de natureza credora" in text:
        return AMARELO_BG
    if "compensacao" in text:
        return VERDE_BG
    return CINZA_LINHA


def group_fill(value: object) -> str:
    text = normalize_text(value)
    if "ativo" in text:
        return VERDE_BG
    if "passivo" in text:
        return VERMELHO_BG
    if "receita" in text:
        return AMARELO_BG
    return CINZA_LINHA


def safe_days(value: object) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 1


def dataframe_to_excel(df: pd.DataFrame) -> bytes:
    df = df.copy()
    if "Dias impactados" not in df.columns:
        df["Dias impactados"] = 1
    df["Dias impactados"] = df["Dias impactados"].map(safe_days)
    df["Saldo final do dia"] = pd.to_numeric(df.get("Saldo final do dia", 0), errors="coerce").fillna(0)

    buffer = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo Executivo"
    ws.sheet_view.showGridLines = False
    build_excel_summary(ws, df)

    detail = wb.create_sheet("Detalhamento")
    detail.sheet_view.showGridLines = False
    build_excel_detail(detail, df)

    by_account = wb.create_sheet("Por Conta")
    by_account.sheet_view.showGridLines = False
    build_excel_by_account(by_account, df)

    wb.save(buffer)
    return buffer.getvalue()


def build_excel_summary(ws, df: pd.DataFrame) -> None:
    widths = {"A": 2, "B": 24, "C": 16, "D": 24, "E": 16, "F": 24, "G": 16, "H": 2}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.merge_cells("B2:G2")
    ws["B2"] = "ANALISE DE SALDOS DIARIOS - SCI"
    ws["B2"].font = Font(name="Arial", bold=True, color=BRANCO, size=16)
    ws["B2"].fill = PatternFill("solid", fgColor=AZUL_ESCURO)
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 38

    ws.merge_cells("B3:G3")
    ws["B3"] = "Inconsistencias identificadas no sistema contabil"
    ws["B3"].font = Font(name="Arial", italic=True, color=AZUL_MED, size=9)
    ws["B3"].fill = PatternFill("solid", fgColor=AZUL_CLARO)
    ws["B3"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 20

    cards = [
        ("B", "C", "Total de Ocorrencias", len(df), "registros", VERMELHO_BG, VERMELHO_FG),
        ("D", "E", "Contas Analisadas", df["Codigo da conta"].nunique() if not df.empty else 0, "contas", AZUL_CLARO, AZUL_MED),
        ("F", "G", "Dias Impactados", int(df["Dias impactados"].sum()) if not df.empty else 0, "dias acumulados", AMARELO_BG, AMARELO_FG),
    ]
    for start_col, end_col, label, value, unit, fill, fg in cards:
        for row, text, size, bold in [(6, label, 9, True), (7, value, 24, True), (8, unit, 8, False)]:
            ws.merge_cells(f"{start_col}{row}:{end_col}{row}")
            cell = ws[f"{start_col}{row}"]
            cell.value = text
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.font = Font(name="Arial", bold=bold, color=fg, size=size)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = borda_media()
        ws.row_dimensions[6].height = 22
        ws.row_dimensions[7].height = 28
        ws.row_dimensions[8].height = 22

    row = 12
    row = add_distribution_section(ws, df, row, "DISTRIBUICAO POR TIPO DE INCONSISTENCIA", "Tipo de inconsistencia")
    row += 2
    add_distribution_section(ws, df, row, "DISTRIBUICAO POR GRUPO CONTABIL", "Grupo", by_group=True)


def add_distribution_section(ws, df: pd.DataFrame, start_row: int, title: str, group_col: str, by_group: bool = False) -> int:
    ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=7)
    cell = ws.cell(start_row, 2, title)
    estilo_cabecalho(cell, AZUL_MED)
    ws.row_dimensions[start_row].height = 22
    headers = ["Tipo" if not by_group else "Grupo", "Ocorrencias", "Dias Impactados", "Maior Saldo (R$)", "Contas", "% Total"]
    for col, header in enumerate(headers, 2):
        estilo_cabecalho(ws.cell(start_row + 1, col, header), AZUL_ESCURO)

    if df.empty or group_col not in df.columns:
        return start_row + 2

    grouped = (
        df.groupby(group_col)
        .agg(
            ocorrencias=("Data", "count"),
            dias=("Dias impactados", "sum"),
            saldo=("Saldo final do dia", lambda s: abs(s).max()),
            contas=("Codigo da conta", "nunique"),
        )
        .sort_values("ocorrencias", ascending=False)
    )
    total = max(len(df), 1)
    row = start_row + 2
    for name, data in grouped.iterrows():
        fill = group_fill(name) if by_group else issue_fill(name)
        values = [name, int(data["ocorrencias"]), int(data["dias"]), float(data["saldo"]), int(data["contas"]), float(data["ocorrencias"]) / total]
        for col, value in enumerate(values, 2):
            cell = ws.cell(row, col, value)
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.border = borda_fina()
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(horizontal="left" if col == 2 else "center", vertical="center")
            if col == 5:
                cell.number_format = "#,##0.00"
            if col == 7:
                cell.number_format = "0.0%"
        row += 1

    totals = ["TOTAL", len(df), int(df["Dias impactados"].sum()), float(abs(df["Saldo final do dia"]).max() if not df.empty else 0), df["Codigo da conta"].nunique(), 1]
    for col, value in enumerate(totals, 2):
        cell = ws.cell(row, col, value)
        cell.fill = PatternFill("solid", fgColor=AZUL_ESCURO)
        cell.font = Font(name="Arial", bold=True, color=BRANCO, size=9)
        cell.border = borda_fina()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if col == 5:
            cell.number_format = "#,##0.00"
        if col == 7:
            cell.number_format = "0.0%"
    return row


def build_excel_detail(ws, df: pd.DataFrame) -> None:
    widths = {"A": 8, "B": 30, "C": 12, "D": 10, "E": 12, "F": 15, "G": 10, "H": 45, "I": 8, "J": 14}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.merge_cells("A1:J1")
    ws["A1"] = "ANALISE DE SALDOS DIARIOS - DETALHAMENTO COMPLETO"
    estilo_cabecalho(ws["A1"], AZUL_ESCURO)
    ws["A1"].font = Font(name="Arial", bold=True, color=BRANCO, size=13)
    ws.row_dimensions[1].height = 36
    headers = ["Codigo", "Conta", "Grupo", "Natureza", "Data", "Saldo (R$)", "Dias Impact.", "Tipo de Inconsistencia", "Seq.", "Dt. Final Seq."]
    for col, header in enumerate(headers, 1):
        estilo_cabecalho(ws.cell(2, col, header), AZUL_MED)

    ordered = df.sort_values(["Grupo", "Codigo da conta", "Data"]) if not df.empty else df
    for row_idx, (_, row) in enumerate(ordered.iterrows(), 3):
        values = [
            row.get("Codigo da conta", ""),
            row.get("Nome da conta no razao", ""),
            row.get("Grupo", ""),
            row.get("Natureza esperada", ""),
            row.get("Data", ""),
            abs(float(row.get("Saldo final do dia", 0))),
            safe_days(row.get("Dias impactados", 1)),
            row.get("Tipo de inconsistencia", ""),
            "sim" if safe_days(row.get("Dias impactados", 1)) > 1 else "",
            row.get("Data final da sequencia", ""),
        ]
        fill = issue_fill(row.get("Tipo de inconsistencia", ""))
        for col, value in enumerate(values, 1):
            cell = ws.cell(row_idx, col, value)
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.border = borda_fina()
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(horizontal="left" if col in {2, 8} else "center", vertical="center", wrap_text=col in {2, 8})
            if col == 6:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[row_idx].height = 17
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:J{max(2, len(ordered) + 2)}"


def build_excel_by_account(ws, df: pd.DataFrame) -> None:
    widths = {"A": 8, "B": 34, "C": 14, "D": 12, "E": 16, "F": 15, "G": 42}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.merge_cells("A1:G1")
    ws["A1"] = "ANALISE DE SALDOS DIARIOS - CONSOLIDADO POR CONTA"
    estilo_cabecalho(ws["A1"], AZUL_ESCURO)
    ws["A1"].font = Font(name="Arial", bold=True, color=BRANCO, size=13)
    ws.row_dimensions[1].height = 36
    headers = ["Codigo", "Conta", "Grupo", "Ocorrencias", "Total Dias Impact.", "Maior Saldo (R$)", "Tipo Principal"]
    for col, header in enumerate(headers, 1):
        estilo_cabecalho(ws.cell(2, col, header), AZUL_MED)
    if df.empty:
        return
    grouped = (
        df.groupby(["Codigo da conta", "Nome da conta no razao", "Grupo"])
        .agg(
            ocorrencias=("Data", "count"),
            dias=("Dias impactados", "sum"),
            saldo=("Saldo final do dia", lambda s: abs(s).max()),
            tipo=("Tipo de inconsistencia", lambda s: s.mode().iloc[0] if not s.mode().empty else ""),
        )
        .reset_index()
        .sort_values("ocorrencias", ascending=False)
    )
    for row_idx, (_, row) in enumerate(grouped.iterrows(), 3):
        values = [row["Codigo da conta"], row["Nome da conta no razao"], row["Grupo"], int(row["ocorrencias"]), int(row["dias"]), float(row["saldo"]), row["tipo"]]
        fill = group_fill(row["Grupo"])
        for col, value in enumerate(values, 1):
            cell = ws.cell(row_idx, col, value)
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.border = borda_fina()
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(horizontal="left" if col in {2, 7} else "center", vertical="center", wrap_text=col in {2, 7})
            if col == 6:
                cell.number_format = "#,##0.00"

